#!/usr/bin/env python3
"""Import Amori dog-owner survey contacts into local CRM and WEEEK.

The script is intentionally one-shot and safe to rerun: imported rows are tagged
with [survey_row=N] in notes and skipped on later executions.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
AGENTS_DIR = REPO_ROOT / "agents"
if str(AGENTS_DIR) not in sys.path:
    sys.path.insert(0, str(AGENTS_DIR))


SURVEY_SOURCE = "survey_2026_07_16"
CONFIRM_TOKEN = "2026-07-16-survey"
TEST_LEAD_IDS = (1, 2, 3)

OWNER_COL = "Являетесь ли вы владельцем собаки?"
VET_COL = "Вы работаете в сфере ветеринарии?"
VET_CONTACT_COL = "Если вы работаете в сфере ветеринарии, оставьте контакт для связи"
PET_COUNT_COL = "Сколько собак у вас сейчас?"
PET_AGE_COL = "Возраст питомца(ев)"
VET_VISIT_COL = "Как часто вы посещаете ветеринара?"
VET_SPEND_COL = "Сколько в среднем вы тратите на ветеринарные услуги в год?"
MAIN_PROBLEM_COL = "Какую проблему в уходе за собакой вы хотели бы решить в первую очередь?"
TRACKING_NOW_COL = "Используете ли вы сейчас какие-либо приложения или устройства для отслеживания здоровья собаки?"
TRACKING_WHAT_COL = "Если да, какие?"
USEFUL_247_COL = "Насколько полезным вам кажется устройство, которое автоматически отслеживает состояние собаки 24/7?"
INTERVIEW_COL = "Готовы ли вы принять участие в коротком интервью (15 минут), чтобы помочь нам лучше понять потребности владельцев собак?"
CONTACT_COL = "Если готовы, оставьте удобный контакт (телефон или Telegram)"

SIZE_PREFIX = "Размер питомца(ев) / "
FOOD_PREFIX = "Как вы кормите вашу собаку? / "
HEALTH_PREFIX = "Какие проблемы со здоровьем питомца(ев) вас беспокоят больше всего? / "
FEATURE_PREFIX = "Какие функции были бы наиболее полезны в приложении для здоровья собаки? / "

EMAIL_RE = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
TELEGRAM_RE = re.compile(r"(?<![\w])@([A-Za-z0-9_]{3,32})")
PHONE_RE = re.compile(r"(?:\+?\d[\d\s().-]{8,}\d)")
SURVEY_ROW_RE = re.compile(r"\[survey_row=(\d+)\]")
CYRILLIC_NAME_RE = re.compile(r"\b[А-ЯЁ][а-яё]+(?:[- ][А-ЯЁ][а-яё]+){0,2}\b")
NAME_STOPWORDS = {
    "telegram",
    "tg",
    "whatsapp",
    "контакт",
    "мессенджер",
    "связь",
    "телефон",
    "телеграм",
    "ватсап",
    "владивосток",
    "владивостоке",
    "москва",
    "москве",
}


@dataclass(frozen=True)
class SurveyLead:
    row_number: int
    name: str
    real_name: str | None
    email: str | None
    phone: str | None
    telegram: str | None
    raw_contact: str
    pet_count: int | None
    notes: str


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def is_yes(value) -> bool:
    return clean(value).lower() == "да"


def parse_pet_count(value) -> int | None:
    text = clean(value)
    match = re.search(r"\d+", text)
    return int(match.group(0)) if match else None


def normalize_phone(raw: str) -> str | None:
    match = PHONE_RE.search(raw)
    if not match:
        return None
    candidate = match.group(0)
    digits = re.sub(r"\D", "", candidate)
    if len(digits) < 10:
        return None
    if len(digits) == 10:
        return "+7" + digits
    if len(digits) == 11 and digits.startswith("8"):
        return "+7" + digits[1:]
    if candidate.strip().startswith("+"):
        return "+" + digits
    return digits


def parse_contact(raw: str) -> tuple[str | None, str | None, str | None]:
    text = clean(raw)
    email_match = EMAIL_RE.search(text)
    tg_match = TELEGRAM_RE.search(text)
    email = email_match.group(0) if email_match else None
    telegram = "@" + tg_match.group(1) if tg_match else None
    phone = normalize_phone(text)
    return email, phone, telegram


def contact_label(email: str | None, phone: str | None, telegram: str | None, raw_contact: str = "") -> str:
    if telegram:
        return telegram
    if phone:
        return phone
    if email:
        return email
    return clean(raw_contact) or "контакт не распознан"


def extract_name_from_contact(raw: str) -> str | None:
    text = clean(raw)
    if not text:
        return None
    text = EMAIL_RE.sub(" ", text)
    text = TELEGRAM_RE.sub(" ", text)
    text = PHONE_RE.sub(" ", text)
    text = re.sub(r"[^\wА-Яа-яЁё -]+", " ", text)
    text = re.sub(
        r"\b(телефон|тел|telegram|tg|тг|whatsapp|ватсап|телеграм|мессенджер|связь|контакт)\b",
        " ",
        text,
        flags=re.I,
    )
    text = re.sub(r"\s+", " ", text).strip()
    for match in CYRILLIC_NAME_RE.finditer(text):
        words = [word.strip("- ") for word in match.group(0).split() if word.strip("- ")]
        normalized = [word.lower() for word in words]
        if len(words) < 2:
            continue
        if any(word in NAME_STOPWORDS for word in normalized):
            continue
        return " ".join(word[:1].upper() + word[1:].lower() for word in words)
    return None


def extract_name_from_telegram(telegram: str | None) -> str | None:
    if not telegram:
        return None
    handle = telegram.lstrip("@")
    if not re.fullmatch(r"[A-Za-z]+_[A-Za-z]+", handle):
        return None
    first, last = handle.split("_", 1)
    if len(first) < 3 or len(last) < 3:
        return None
    if first.lower() in NAME_STOPWORDS or last.lower() in NAME_STOPWORDS:
        return None
    return f"{first.title()} {last.title()}"


def lead_display_name(row_number: int, raw_contact: str, email: str | None, phone: str | None, telegram: str | None) -> tuple[str, str | None]:
    real_name = extract_name_from_contact(raw_contact) or extract_name_from_telegram(telegram)
    if real_name:
        return real_name, real_name
    return f"Респондент анкеты #{row_number}", None


def selected_options(headers: list[str], row: tuple, prefix: str) -> list[str]:
    items = []
    for idx, header in enumerate(headers):
        if not header.startswith(prefix):
            continue
        value = clean(row[idx] if idx < len(row) else None)
        if not value:
            continue
        option = header.removeprefix(prefix)
        if value.lower() in {"да", "yes", "true", "1", option.lower()}:
            items.append(option)
        else:
            items.append(f"{option}: {value}")
    return items


def get_by_header(headers: list[str], row: tuple, name: str) -> str:
    try:
        idx = headers.index(name)
    except ValueError:
        return ""
    return clean(row[idx] if idx < len(row) else None)


def build_notes(headers: list[str], row: tuple, row_number: int, raw_contact: str) -> str:
    sizes = selected_options(headers, row, SIZE_PREFIX)
    foods = selected_options(headers, row, FOOD_PREFIX)
    health = selected_options(headers, row, HEALTH_PREFIX)
    features = selected_options(headers, row, FEATURE_PREFIX)

    parts = [
        f"[survey_row={row_number}]",
        f"Источник: анкета владельцев собак от 16.07.2026",
        f"Владелец собаки: {get_by_header(headers, row, OWNER_COL) or 'не указано'}",
        f"Работает в ветеринарии: {get_by_header(headers, row, VET_COL) or 'не указано'}",
        f"Количество собак: {get_by_header(headers, row, PET_COUNT_COL) or 'не указано'}",
        f"Размеры: {', '.join(sizes) if sizes else 'не указано'}",
        f"Возраст: {get_by_header(headers, row, PET_AGE_COL) or 'не указано'}",
        f"Кормление: {', '.join(foods) if foods else 'не указано'}",
        f"Ветеринар: {get_by_header(headers, row, VET_VISIT_COL) or 'не указано'}",
        f"Ветрасходы в год: {get_by_header(headers, row, VET_SPEND_COL) or 'не указано'}",
        f"Боли по здоровью: {', '.join(health) if health else 'не указано'}",
        f"Главная проблема: {get_by_header(headers, row, MAIN_PROBLEM_COL) or 'не указано'}",
        f"Уже использует трекеры: {get_by_header(headers, row, TRACKING_NOW_COL) or 'не указано'}",
        f"Какие трекеры: {get_by_header(headers, row, TRACKING_WHAT_COL) or 'не указано'}",
        f"Полезные функции: {', '.join(features) if features else 'не указано'}",
        f"Оценка 24/7 устройства: {get_by_header(headers, row, USEFUL_247_COL) or 'не указано'}",
        f"Готов к интервью: {get_by_header(headers, row, INTERVIEW_COL) or 'не указано'}",
        f"Raw contact: {raw_contact}",
    ]

    vet_contact = get_by_header(headers, row, VET_CONTACT_COL)
    if vet_contact:
        parts.append(f"Контакт из ветблока: {vet_contact}")
    return "\n".join(parts)


def build_weeek_description(lead: SurveyLead) -> str:
    contact = contact_label(lead.email, lead.phone, lead.telegram, lead.raw_contact)
    title_name = lead.real_name or lead.name
    contact_lines = [
        "КАРТОЧКА ЛИДА AMORI",
        "",
        f"Имя/ориентир: {title_name}",
        f"Источник: анкета владельцев собак 16.07.2026, строка #{lead.row_number}",
        f"Тип: B2C, питомец: собака",
        "",
        "КАК СВЯЗАТЬСЯ",
        f"Основной контакт: {contact}",
    ]
    if lead.telegram:
        contact_lines.append(f"Telegram: {lead.telegram}")
    if lead.phone:
        contact_lines.append(f"Телефон: {lead.phone}")
    if lead.email:
        contact_lines.append(f"Email: {lead.email}")
    contact_lines.extend(
        [
            f"Raw contact из анкеты: {lead.raw_contact or 'не указано'}",
            "",
            "ОТВЕТЫ АНКЕТЫ",
            lead.notes,
        ]
    )
    return "\n".join(contact_lines)


def build_weeek_title(lead: SurveyLead) -> str:
    contact = contact_label(lead.email, lead.phone, lead.telegram, lead.raw_contact)
    if lead.real_name:
        title = f"{lead.real_name} — анкета Amori #{lead.row_number}"
    else:
        title = f"Анкета Amori #{lead.row_number} — {contact}"
    return title[:180]


def iter_nonempty_rows(path: str | Path) -> Iterable[tuple[int, tuple]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(clean(cell) for cell in row):
            yield row_number, row


def read_survey_leads(path: str | Path) -> list[SurveyLead]:
    rows = list(iter_nonempty_rows(path))
    if not rows:
        return []
    headers = [clean(v) for v in rows[0][1]]
    leads: list[SurveyLead] = []
    for row_number, row in rows[1:]:
        contact = get_by_header(headers, row, CONTACT_COL)
        if not is_yes(get_by_header(headers, row, INTERVIEW_COL)) or not contact:
            continue
        email, phone, telegram = parse_contact(contact)
        name, real_name = lead_display_name(row_number, contact, email, phone, telegram)
        leads.append(
            SurveyLead(
                row_number=row_number,
                name=name,
                real_name=real_name,
                email=email,
                phone=phone,
                telegram=telegram,
                raw_contact=contact,
                pet_count=parse_pet_count(get_by_header(headers, row, PET_COUNT_COL)),
                notes=build_notes(headers, row, row_number, contact),
            )
        )
    return leads


def imported_rows(cur) -> set[int]:
    cur.execute(
        "SELECT notes FROM leads WHERE source=%s AND notes LIKE %s",
        (SURVEY_SOURCE, "[survey_row=%"),
    )
    rows = set()
    for (notes,) in cur.fetchall():
        match = SURVEY_ROW_RE.search(notes or "")
        if match:
            rows.add(int(match.group(1)))
    return rows


def current_test_leads(cur) -> list[tuple]:
    cur.execute(
        """
        SELECT id, name, weeek_deal_id, weeek_contact_id
        FROM leads
        WHERE id = ANY(%s)
        ORDER BY id
        """,
        (list(TEST_LEAD_IDS),),
    )
    return cur.fetchall()


def update_imported_lead(lead_id: int, pet_count: int | None):
    import db

    conn = db.connect("customer_db")
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE leads
        SET interest_level='warm',
            status='new',
            pet_count=COALESCE(%s, pet_count),
            next_followup_at=NOW() + INTERVAL '1 day',
            updated_at=NOW()
        WHERE id=%s
        """,
        (pet_count, lead_id),
    )
    conn.commit()
    cur.close()
    conn.close()


def update_existing_imported_lead(lead_id: int, lead: SurveyLead) -> None:
    import db

    conn = db.connect("customer_db")
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE leads
        SET name=%s,
            email=COALESCE(%s, email),
            phone=COALESCE(%s, phone),
            telegram_username=COALESCE(%s, telegram_username),
            notes=%s,
            interest_level='warm',
            status='new',
            pet_count=COALESCE(%s, pet_count),
            next_followup_at=COALESCE(next_followup_at, NOW() + INTERVAL '1 day'),
            updated_at=NOW()
        WHERE id=%s
        """,
        (lead.name, lead.email, lead.phone, lead.telegram, lead.notes, lead.pet_count, lead_id),
    )
    conn.commit()
    cur.close()
    conn.close()


def sync_missing_weeek_for_source() -> dict:
    import db
    from lead_manager import create_weeek_contact, create_weeek_deal

    conn = db.connect("customer_db")
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, name, email, phone, pet_type, weeek_contact_id, weeek_deal_id
        FROM leads
        WHERE source=%s
          AND (weeek_contact_id IS NULL OR weeek_deal_id IS NULL)
        ORDER BY id
        """,
        (SURVEY_SOURCE,),
    )
    rows = cur.fetchall()

    contacts_created = 0
    deals_created = 0
    still_missing = 0
    for lead_id, name, email, phone, pet_type, contact_id, deal_id in rows:
        if not contact_id:
            contact_id = create_weeek_contact(name, email, phone)
            if contact_id:
                contacts_created += 1
                cur.execute(
                    "UPDATE leads SET weeek_contact_id=%s, updated_at=NOW() WHERE id=%s",
                    (contact_id, lead_id),
                )
        if contact_id and not deal_id:
            deal_id = create_weeek_deal(
                f"Лид: {name}" + (f" — {pet_type}" if pet_type else ""),
                contact_id,
                "new",
            )
            if deal_id:
                deals_created += 1
                cur.execute(
                    "UPDATE leads SET weeek_deal_id=%s, updated_at=NOW() WHERE id=%s",
                    (deal_id, lead_id),
                )
        if not contact_id or not deal_id:
            still_missing += 1

    conn.commit()
    cur.close()
    conn.close()
    return {
        "checked": len(rows),
        "contacts_created": contacts_created,
        "deals_created": deals_created,
        "still_missing": still_missing,
    }


def sync_existing_survey_to_weeek(path: str | Path, confirm: str) -> int:
    if confirm != CONFIRM_TOKEN:
        print(f"Refusing to sync: pass --confirm-import={CONFIRM_TOKEN}", file=sys.stderr)
        return 2

    import db
    from lead_manager import (
        create_weeek_contact,
        create_weeek_deal,
        upsert_weeek_deal_task,
        update_weeek_contact,
        update_weeek_deal_details,
    )

    leads_by_row = {lead.row_number: lead for lead in read_survey_leads(path)}
    conn = db.connect("customer_db")
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, notes, weeek_contact_id, weeek_deal_id
        FROM leads
        WHERE source=%s
        ORDER BY id
        """,
        (SURVEY_SOURCE,),
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    matched = 0
    local_updated = 0
    contacts_created = 0
    contacts_updated = 0
    deals_created = 0
    deals_updated = 0
    deal_tasks_created = 0
    deal_tasks_updated = 0
    weeek_warnings = 0

    for lead_id, notes, contact_id, deal_id in rows:
        match = SURVEY_ROW_RE.search(notes or "")
        if not match:
            continue
        lead = leads_by_row.get(int(match.group(1)))
        if not lead:
            continue
        matched += 1
        update_existing_imported_lead(int(lead_id), lead)
        local_updated += 1

        if not contact_id:
            contact_id = create_weeek_contact(lead.name, lead.email, lead.phone)
            if contact_id:
                contacts_created += 1
                conn = db.connect("customer_db")
                cur = conn.cursor()
                cur.execute(
                    "UPDATE leads SET weeek_contact_id=%s, updated_at=NOW() WHERE id=%s",
                    (contact_id, lead_id),
                )
                conn.commit()
                cur.close()
                conn.close()
        elif update_weeek_contact(contact_id, lead.name, lead.email, lead.phone).get("ok"):
            contacts_updated += 1
        else:
            weeek_warnings += 1

        description = build_weeek_description(lead)
        title = build_weeek_title(lead)
        if contact_id and not deal_id:
            deal_id = create_weeek_deal(title, contact_id, "new", description=description)
            if deal_id:
                deals_created += 1
                conn = db.connect("customer_db")
                cur = conn.cursor()
                cur.execute(
                    "UPDATE leads SET weeek_deal_id=%s, updated_at=NOW() WHERE id=%s",
                    (deal_id, lead_id),
                )
                conn.commit()
                cur.close()
                conn.close()
        elif deal_id:
            result = update_weeek_deal_details(deal_id, title=title, description=description, stage="new")
            if result.get("ok"):
                deals_updated += 1
            else:
                weeek_warnings += 1
        if deal_id:
            task_result = upsert_weeek_deal_task(
                deal_id,
                "Анкета лида Amori: ответы и контакт",
                description,
            )
            if task_result.get("ok") and task_result.get("action") == "created":
                deal_tasks_created += 1
            elif task_result.get("ok") and task_result.get("action") in {"updated", "replaced", "kept_existing"}:
                deal_tasks_updated += 1
            else:
                weeek_warnings += 1

    print(f"survey_source={SURVEY_SOURCE}")
    print(f"matched_existing={matched}")
    print(f"local_updated={local_updated}")
    print(f"weeek_contacts_created={contacts_created}")
    print(f"weeek_contacts_updated={contacts_updated}")
    print(f"weeek_deals_created={deals_created}")
    print(f"weeek_deals_updated={deals_updated}")
    print(f"weeek_deal_tasks_created={deal_tasks_created}")
    print(f"weeek_deal_tasks_updated={deal_tasks_updated}")
    print(f"weeek_warnings={weeek_warnings}")
    return 0 if weeek_warnings == 0 else 1


def dry_run(path: str | Path) -> int:
    import db

    leads = read_survey_leads(path)
    conn = db.connect("customer_db")
    cur = conn.cursor()
    existing_rows = imported_rows(cur)
    tests = current_test_leads(cur)
    cur.close()
    conn.close()

    to_create = [lead for lead in leads if lead.row_number not in existing_rows]
    print(f"survey_source={SURVEY_SOURCE}")
    print(f"contact_leads_in_file={len(leads)}")
    print(f"already_imported={len(existing_rows)}")
    print(f"to_create={len(to_create)}")
    print(f"test_leads_to_remove={len(tests)}")
    print("dry_run=true")
    return 0


def execute(path: str | Path, confirm: str) -> int:
    if confirm != CONFIRM_TOKEN:
        print(f"Refusing to execute: pass --confirm-import={CONFIRM_TOKEN}", file=sys.stderr)
        return 2

    import db
    from lead_manager import add_lead, remove_test_lead, update_weeek_deal_details, upsert_weeek_deal_task

    leads = read_survey_leads(path)
    conn = db.connect("customer_db")
    cur = conn.cursor()
    existing_rows = imported_rows(cur)
    tests = current_test_leads(cur)
    cur.close()
    conn.close()

    removed = []
    for lead_id, *_ in tests:
        removed.append(remove_test_lead(int(lead_id)))

    created = []
    skipped = 0
    enriched_weeek_deals = 0
    for lead in leads:
        if lead.row_number in existing_rows:
            skipped += 1
            continue
        result = add_lead(
            name=lead.name,
            email=lead.email,
            phone=lead.phone,
            telegram=lead.telegram,
            source=SURVEY_SOURCE,
            pet_type="собака",
            notes=lead.notes,
            lead_type="b2c",
        )
        update_imported_lead(result["id"], lead.pet_count)
        if result.get("weeek_deal_id"):
            description = build_weeek_description(lead)
            title = build_weeek_title(lead)
            update_weeek_deal_details(result["weeek_deal_id"], title=title, description=description, stage="new")
            task_result = upsert_weeek_deal_task(
                result["weeek_deal_id"],
                "Анкета лида Amori: ответы и контакт",
                description,
            )
            if task_result.get("ok"):
                enriched_weeek_deals += 1
        created.append(result)

    sync = sync_missing_weeek_for_source()
    weeek_contacts = sum(1 for item in created if item.get("weeek_contact_id"))
    weeek_deals = sum(1 for item in created if item.get("weeek_deal_id"))
    weeek_remove_warnings = sum(
        1
        for item in removed
        for result in (item.get("weeek") or {}).values()
        if not result.get("ok")
    )

    print(f"survey_source={SURVEY_SOURCE}")
    print(f"created_local={len(created)}")
    print(f"created_weeek_contacts={weeek_contacts}")
    print(f"created_weeek_deals={weeek_deals}")
    print(f"enriched_weeek_deals={enriched_weeek_deals}")
    print(f"skipped_existing={skipped}")
    print(f"synced_missing_weeek_checked={sync['checked']}")
    print(f"synced_missing_weeek_contacts={sync['contacts_created']}")
    print(f"synced_missing_weeek_deals={sync['deals_created']}")
    print(f"synced_missing_weeek_still_missing={sync['still_missing']}")
    print(f"removed_test_leads={sum(1 for item in removed if item.get('deleted'))}")
    print(f"weeek_remove_warnings={weeek_remove_warnings}")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, help="Path to XLSX survey export")
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument("--dry-run", action="store_true", help="Show planned changes only")
    mode.add_argument("--execute", action="store_true", help="Import leads and remove test data")
    mode.add_argument("--sync-existing", action="store_true", help="Refresh already imported survey leads in local CRM and WEEEK")
    parser.add_argument("--confirm-import", default="", help="Required for --execute")
    args = parser.parse_args(argv)

    path = Path(args.file).expanduser()
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 1
    if args.dry_run:
        return dry_run(path)
    if args.sync_existing:
        return sync_existing_survey_to_weeek(path, args.confirm_import)
    return execute(path, args.confirm_import)


if __name__ == "__main__":
    raise SystemExit(main())
